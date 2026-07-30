"""
Reads Lifting_Consolidated_AllYears.xlsx (sibling "Lifting Data" project) and
writes ../data/lifting.json for the 4-tab Lifting & Strength dashboard:
leaderboards split by class year or by All-Time/Last-Session, plus per-athlete
time series for the Class Comparison tab.

Class year (Freshman/Sophomore/Junior/Senior) isn't its own column in the Lifting
Data output -- `years_with_program` is used as the proxy (confirmed identical to
the source's own CLASS field for every athlete checked in that project's own
validation, see Lifting Data/.claude/skills/lifting-data/SKILL.md). Only populated
2023-24 onward, so 2021-22/2022-23 athletes only appear in the All-Time/Last-Session
tab, never a class-specific one.

Extended 2026-07-30 (per the user, "reporting dashboards of the other metrics"):
beyond the original 4 lift metrics, the Lifting Data pipeline also computes two
composite z-score-based percentiles (Strength Score, Athleticism Score -- see
that project's SKILL.md "Scores" section) and tracks three raw athletic-testing
metrics (Broad Jump, Vertical, Pro Agility) that this dashboard never surfaced
before. Team-scope scores only (not Position-scope) for this first pass -- see
this project's README for why. Athleticism Score only exists for 2023-24/2025-26
(its own "require all 6 metrics" gate -- Vertical/Pro Agility don't co-exist in
every year); Pro Agility itself is entirely absent from 2024-25. Both are real
source-data gaps carried straight through, not bugs to paper over.

Re-run whenever the Lifting Data output changes:
    python build_lifting_data.py
"""

import json
from pathlib import Path

import openpyxl

SRC = Path(__file__).resolve().parent.parent.parent / "Lifting Data" / "output" / "Lifting_Consolidated_AllYears.xlsx"
OUT = Path(__file__).resolve().parent.parent / "data" / "lifting.json"

PERIOD_ORDER = {"December": 0, "January": 1, "April": 2}
CLASS_NAMES = {1: "Freshman", 2: "Sophomore", 3: "Junior", 4: "Senior"}

STRENGTH_METRICS = ["Combined Total", "Bench", "Squat", "Clean", "Strength Score (Team)"]
ATHLETIC_METRICS = ["Broad Jump", "Vertical", "Pro Agility", "Athleticism Score (Team)"]
LEADERBOARD_METRICS = STRENGTH_METRICS + ATHLETIC_METRICS
# Pro Agility is a timed sprint -- lower is faster/better, the one metric on this
# page that isn't "bigger number wins" (matches Lifting Data's own LOWER_IS_BETTER
# convention for its z-score math).
LOWER_IS_BETTER = {"Pro Agility"}
# Metrics whose real value is `value` directly (a percentile, or -- for Broad
# Jump/Vertical/Pro Agility -- the specific attempt='Best' row's own value) rather
# than `calculated_1rm` (the lift 1RM the original 4 metrics use).
SCORE_METRICS = {"Strength Score (Team)", "Athleticism Score (Team)"}
BEST_ATTEMPT_METRICS = {"Broad Jump", "Vertical", "Pro Agility"}


def metric_value(r):
    """The real per-row value for whichever of the LEADERBOARD_METRICS this row
    is, or None if this specific row isn't the one to use (e.g. a Broad Jump
    attempt=1/2 row, which exists alongside the attempt='Best' row this reads
    instead -- using every attempt row too would just redundantly re-compare the
    same session's already-known best against itself)."""
    metric = r["metric"]
    if metric in SCORE_METRICS:
        return r["value"]
    if metric in BEST_ATTEMPT_METRICS:
        return r["value"] if r["attempt"] == "Best" else None
    return r["calculated_1rm"]  # lifts + Combined Total


def is_better(metric, v, cur):
    """True if `v` should replace `cur` as the best-seen value for `metric`."""
    return v < cur if metric in LOWER_IS_BETTER else v > cur


def session_key(football_year, period):
    start_year = int(football_year[:4])
    return (start_year, PERIOD_ORDER.get(period, 9))


def session_label(football_year, period):
    return f"{period} {football_year}"


def class_name(years_with_program):
    if years_with_program is None:
        return None
    yrs = int(years_with_program)
    return CLASS_NAMES.get(yrs, f"{yrs}th Year" if yrs >= 5 else None)


def main():
    wb = openpyxl.load_workbook(SRC)
    ws = wb["Data"]
    headers = [c.value for c in ws[1]]
    rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]

    sessions = sorted({(r["football_year"], r["testing_period"]) for r in rows if r["testing_period"]},
                       key=lambda s: session_key(*s))
    last_session = sessions[-1] if sessions else None

    # best value per (athlete_key, football_year, testing_period, metric) --
    # Combined Total is already one row; Bench/Squat/Clean take the max of
    # whatever attempts exist that session (same "best" logic the Lifting Data
    # project itself uses for Combined Total, so leaderboard values agree with it).
    #
    # meta is built from EVERY row, not just LEADERBOARD_METRICS ones -- originally
    # only populated inside the leaderboard-metric loop below, which left
    # (athlete_key, football_year) pairs with a name of None whenever an athlete's
    # only data that year was Height/Weight (no lift attempt at all that season).
    # Real bug found 2026-07-29 testing in the browser: 12 "null null" entries in
    # the Class Comparison athlete dropdown.
    meta = {}
    for r in rows:
        if not r["athlete_key"]:
            continue
        key = (r["athlete_key"], r["football_year"])
        if key not in meta or r["years_with_program"] is not None:
            meta[key] = {
                "first_name": r["first_name"], "last_name": r["last_name"],
                "position": r["position"], "years_with_program": r["years_with_program"],
                "class": class_name(r["years_with_program"]),
            }

    best = {}
    for r in rows:
        metric = r["metric"]
        if metric not in LEADERBOARD_METRICS:
            continue
        if r["testing_period"] is None:
            continue
        v = metric_value(r)
        if v is None:
            continue
        key = (r["athlete_key"], r["football_year"], r["testing_period"], metric)
        if key not in best or is_better(metric, v, best[key]):
            best[key] = v

    leaderboard_rows = []
    for (athlete_key, fy, period, metric), value in best.items():
        m = meta.get((athlete_key, fy), {})
        leaderboard_rows.append({
            "athlete_key": athlete_key, "first_name": m.get("first_name"),
            "last_name": m.get("last_name"), "position": m.get("position"),
            "class": m.get("class"), "football_year": fy, "testing_period": period,
            "session_label": session_label(fy, period), "is_last_session": (fy, period) == last_session,
            "metric": metric, "value": round(value, 1),
        })

    # per-athlete time series (Height/Weight + best value per session for every
    # leaderboard metric), sorted chronologically, for the Class Comparison tab's
    # line charts.
    series_metrics = LEADERBOARD_METRICS + ["Height", "Weight"]
    series_points = {}
    for r in rows:
        metric = r["metric"]
        if metric not in series_metrics:
            continue
        ak = r["athlete_key"]
        if not ak:
            continue
        if metric in ("Height", "Weight"):
            v = r["value"]
        else:
            v = metric_value(r)
        if v is None:
            continue
        fy = r["football_year"]
        period = r["testing_period"] or "—"
        key = (ak, fy, period, metric)
        if metric == "Height":
            # Height has no testing_period (recorded once/session-independent) --
            # collapse to one point per football_year so it still lines up on the
            # same x-axis as the other per-session metrics.
            key = (ak, fy, "—", metric)
        if key not in series_points or is_better(metric, v, series_points[key]):
            series_points[key] = v

    athletes = {}
    for (ak, fy, period, metric), v in series_points.items():
        m = meta.get((ak, fy), {})
        athletes.setdefault(ak, {
            "athlete_key": ak, "first_name": m.get("first_name"), "last_name": m.get("last_name"),
            "points": [],
        })
        athletes[ak]["points"].append({
            "football_year": fy, "testing_period": period,
            "session_label": session_label(fy, period) if period != "—" else fy,
            "metric": metric, "value": round(v, 1),
        })
    for a in athletes.values():
        a["points"].sort(key=lambda p: session_key(p["football_year"], p["testing_period"] if p["testing_period"] != "—" else "December"))

    payload = {
        "generated_from": SRC.name,
        "sessions": [{"football_year": fy, "testing_period": p, "label": session_label(fy, p)} for fy, p in sessions],
        "last_session": {"football_year": last_session[0], "testing_period": last_session[1],
                          "label": session_label(*last_session)} if last_session else None,
        "leaderboard_rows": leaderboard_rows,
        "athletes": list(athletes.values()),
    }

    OUT.parent.mkdir(exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=0, default=str)

    print(f"leaderboard rows: {len(leaderboard_rows)}")
    print(f"athletes with series data: {len(athletes)}")
    print(f"last session: {payload['last_session']['label']}")


if __name__ == "__main__":
    main()
