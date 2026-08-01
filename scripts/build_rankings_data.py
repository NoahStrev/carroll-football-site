"""
Phase 4, first report: reads the already-computed output of two sibling
projects -- CCIW Buddah Report (cciw.org, conference-only ranks) and National
Buddah Report (NCAA.com D3, national ranks) -- and writes ../data/rankings.json
for the new Rankings dashboard (Overall / Offensive / Defensive tabs, CCIW and
National shown side by side within each).

2026-07-31: the "Overall" tab is displayed as "Additional Metrics" in
rankings.html (per the user -- Turnover Margin/Winning %/Penalties read more
like a catch-all than a team's single most important summary). The internal
phase key below is still "overall" throughout this script and the JSON
payload -- only the on-page label changed, not the data shape.

Per the user (2026-07-30): "i have a report that will be repopulated weekly
with the cciw rankings if they're playing a conference opponent and a
national one for whoever they play, i would like an overall section, an
offensive section, and a defensive section". This first pass does NOT
re-scrape either source live -- it reads the two sibling projects' own
already-scraped/classified workbooks, matching the user's own "since the data
is already there make some updates so those reports are appearing" framing
and "site ingestion only for now" scope decision (live weekly automation of
the report-generation step itself is a deferred follow-up, not built here).

Team-level only for this first pass (no individual-leader rows) -- "Overall"
metrics (Turnover Margin, Winning %, Penalties) have no individual-player
equivalent, so scoping the whole page to team-level keeps Overall/Offense/
Defense symmetric instead of Offense/Defense having an extra sub-section
Overall can't have.

Real, permanent asymmetry between the two sources, not a bug: CCIW Buddah
Report's own classification (see that project's classification-rules.md)
never produces an "Overall" phase at all -- cciw.org doesn't publish
Turnover Margin/Win%/Penalties as their own ranked category the way NCAA.com
does, so CCIW-scope data literally cannot populate the Overall tab. This is
called out explicitly in the dashboard rather than silently leaving it empty.

Sources (read directly, not copied into this project):
    ../CCIW Buddah Report/output_carroll/Carroll_Football_AllTime.xlsx
      -- sheets Offense-Team/Defense-Team/SpecialTeams-Team, columns
         Year/Category/Metric/Value/Rank/Out Of. Phase = sheet name directly
         (already classified upstream, see that project's own rules doc).
    ../National Buddah Report/CCIW_D3_Football_Stats/Carroll/Carroll_AllYears.xlsx
      -- sheet "Team Stats", columns Season/Category/Games/Stat/Value/Rank/
         RawRank/Context. Phase is NOT a column here (Category is a free-text
         NCAA.com stat name) -- classified below via CATEGORY_SECTION, ported
         from National Game Prep Report's pull_matchup_data.ps1 (the same
         50-metric Section mapping already proven correct by eye against a
         real generated PDF), plus a handful of near-duplicate categories
         that script doesn't map (e.g. "Fewest Penalties" vs its own
         "Fewest Penalties Per Game") extended by the same judgment-call
         pattern, documented inline below.

2026-07-30 additions:
- Down-to-the-week granularity: cciw_rows/national_rows now keep every in-season
  week's row (not just the latest), tagged with a "week" number, so the dashboard
  can offer a week selector once weekly scraping starts in September. See
  "weeks_by_season" in the payload for what weeks exist per season.
- "This Week" tab: reads ../Schedule/schedule.json (a new sibling project,
  scrape_schedule.py, scraping gopios.com's own schedule page -- no schedule data
  existed anywhere in the workspace before this) to find Carroll's next unplayed
  game, then pulls BOTH Carroll's and that opponent's current CCIW/National rank
  standing for a side-by-side comparison. Opponent data comes from the multi-team
  sibling workbooks (CCIW_TopPerformer_AllTime.xlsx, CCIW/CCIW_AllYears.xlsx) via
  team_cciw_rows()/team_national_rows() -- separate from build_cciw()/
  build_national() above, which stay Carroll-only and untouched. Non-conference
  opponents (not tracked by either Buddah Report project) get
  opponent_data_available: false rather than a guessed/empty table.

2026-07-31 (later): the "This Week" tab described above was replaced in
rankings.html by an "Opponent Scouting" tab built from a different data
source entirely (Offense/Defense's own play-by-play, data/game-data.json,
not this script's output) -- This Week's rank tables just repeated each
phase tab's Carroll row with the opponent's row bolted on, not real new
information. build_this_week() and the opponent-lookup helpers it alone
used (team_cciw_rows/team_national_rows/latest_snapshot_for_season/
_latest_per_group/_resolve_current, plus the CCIW_TOPPERFORMER_SRC/
NATIONAL_CCIW_SRC/SCHEDULE_SRC constants) were removed from this script --
build_cciw()/build_national() and the Carroll-only cciw_rows/national_rows
payload they produce are untouched and still exactly what rankings.html's
phase tabs read.

2026-07-31: dropped the raw-.xlsx "download" feature this script used to build
(copying source workbooks into data/downloads/ and listing them in the payload)
-- per the user, downloads are now a PDF of whatever's currently on screen,
generated client-side via the browser's print dialog (see rankings.html's
printCard()/printPage()), not a pre-built file this script produces. Also, per
the user: a rank number only means something within its own season's real
competitive pool -- rankings.html's "All seasons" view lists each season's real
value but deliberately drops Rank/sorts-by-rank for that view rather than
implying a cross-season comparison neither cciw.org nor NCAA.com actually
publishes; a single selected season still shows its own real rank as published.

Re-run whenever either source project's output changes:
    python build_rankings_data.py
"""

import json
import re
from pathlib import Path

import openpyxl

FOOTBALL_ROOT = Path(__file__).resolve().parent.parent.parent
CCIW_SRC = FOOTBALL_ROOT / "CCIW Buddah Report" / "output_carroll" / "Carroll_Football_AllTime.xlsx"
NATIONAL_SRC = FOOTBALL_ROOT / "National Buddah Report" / "CCIW_D3_Football_Stats" / "Carroll" / "Carroll_AllYears.xlsx"
OUT = Path(__file__).resolve().parent.parent / "data" / "rankings.json"

# In-season weekly progression workbooks -- both sibling projects' weekly-scrape
# scheduled tasks (cciw-org-conference-weekly-scrape, ncaa-d3-national-weekly-scrape)
# write these separately from the AllTime/AllYears files above, which only get rebuilt
# once a season is over ("only on explicit request", per each project's own rule). So
# during the season, a year's rankings live here first, not in AllTime/AllYears yet.
# Picked up automatically -- no per-season path to update by hand each year. Every
# week's row is kept (not collapsed to the latest) -- per the user (2026-07-30), the
# Rankings page needs to go "down to the individual week" once weekly scraping starts,
# so the dashboard can offer a week selector, not just the most recent snapshot.
WEEKLY_YEAR_RE = re.compile(r"_(\d{4})_Weekly\.xlsx$")

CCIW_SHEET_TO_PHASE = {
    "Offense-Team": "offense",
    "Defense-Team": "defense",
    "SpecialTeams-Team": "special_teams",
}

# Ported verbatim from National Game Prep Report/pull_matchup_data.ps1's
# $Metrics array (TeamCat -> Section), the same 50-metric mapping already
# proven correct against a real generated PDF (Weekly Scouting Report). Only
# the TeamCat/Section columns are needed here -- Label/TeamId/IndivCat/IndivId
# are that project's own concern (individual leaders, live NCAA.com fetch IDs).
CATEGORY_SECTION = {
    "Scoring Offense": "offense", "Total Offense": "offense", "Rushing Offense": "offense",
    "Passing Offense": "offense", "Team Passing Efficiency": "offense", "Completion Percentage": "offense",
    "Passing Yards per Completion": "offense", "3rd Down Conversion Pct": "offense",
    "4th Down Conversion Pct": "offense", "Red Zone Offense": "offense", "First Downs Offense": "offense",
    "Time of Possession": "offense", "Turnovers Lost": "offense", "Sacks Allowed": "offense",
    "Tackles for Loss Allowed": "offense",
    "Scoring Defense": "defense", "Total Defense": "defense", "Rushing Defense": "defense",
    "Passing Yards Allowed": "defense", "Team Passing Efficiency Defense": "defense",
    "3rd Down Conversion Pct Defense": "defense", "4th Down Conversion Pct Defense": "defense",
    "Red Zone Defense": "defense", "First Downs Defense": "defense", "Team Sacks": "defense",
    "Team Tackles for Loss": "defense", "Defensive TDs": "defense", "Turnovers Gained": "defense",
    "Net Punting": "special_teams", "Punt Returns": "special_teams", "Kickoff Returns": "special_teams",
    "Punt Return Defense": "special_teams", "Kickoff Return Defense": "special_teams",
    "Blocked Kicks": "special_teams", "Blocked Kicks Allowed": "special_teams",
    "Blocked Punts": "special_teams", "Blocked Punts Allowed": "special_teams",
    "Turnover Margin": "overall", "Winning Percentage": "overall",
    "Fewest Penalties Per Game": "overall", "Fewest Penalty Yards Per Game": "overall",
    # Not in the curated 50-metric list (that script only kept the per-game
    # rate versions) -- same category family, same judgment call extended by
    # this project rather than left unclassified. See module docstring.
    "Fewest Penalties": "overall", "Fewest Penalty Yards": "overall",
    "Fumbles Lost": "offense",       # Carroll's own fumbles lost -- same bucket as "Turnovers Lost"
    "Fumbles Recovered": "defense",  # Carroll recovering a fumble -- same bucket as "Turnovers Gained"
    "Passes Had Intercepted": "offense",  # Carroll QB's own picks thrown -- same bucket as "Turnovers Lost"
    "Passes Intercepted": "defense",      # Carroll defense picking off the opponent -- same bucket as "Turnovers Gained"
}


def sheet_rows(ws):
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for r in ws.iter_rows(min_row=2, values_only=True):
        yield dict(zip(headers, r))


def build_cciw():
    wb = openpyxl.load_workbook(CCIW_SRC, read_only=True, data_only=True)
    rows = []
    for sheet_name, phase in CCIW_SHEET_TO_PHASE.items():
        for r in sheet_rows(wb[sheet_name]):
            if r["Rank"] is None or r["Out Of"] is None:
                continue
            rows.append({
                "phase": phase, "season": str(r["Year"]), "category": r["Category"], "metric": r["Metric"],
                "value": r["Value"], "rank": r["Rank"], "out_of": r["Out Of"], "week": None,
            })
    wb.close()

    finalized_seasons = {r["season"] for r in rows}
    weekly_files = []
    weeks_by_season = {}
    for path in sorted(CCIW_SRC.parent.glob("Carroll_Football_*_Weekly.xlsx")):
        m = WEEKLY_YEAR_RE.search(path.name)
        season = m.group(1) if m else None
        if season is None or season in finalized_seasons:
            continue  # already superseded by that season's finalized AllTime rows
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        season_weeks = {}
        for sheet_name, phase in CCIW_SHEET_TO_PHASE.items():
            if sheet_name not in wb.sheetnames:
                continue
            for r in sheet_rows(wb[sheet_name]):
                if r["Rank"] is None or r["Out Of"] is None:
                    continue
                rows.append({
                    "phase": phase, "season": season, "category": r["Category"], "metric": r["Metric"],
                    "value": r["Value"], "rank": r["Rank"], "out_of": r["Out Of"], "week": r["Week"],
                })
                season_weeks[r["Week"]] = str(r["SnapshotDate"])[:10]
        wb.close()
        if season_weeks:
            weeks_by_season[season] = [{"week": w, "date": season_weeks[w]} for w in sorted(season_weeks)]
            weekly_files.append(path.name)

    return rows, weekly_files, weeks_by_season


def build_national():
    wb = openpyxl.load_workbook(NATIONAL_SRC, read_only=True, data_only=True)
    rows = []
    unmapped = set()
    for r in sheet_rows(wb["Team Stats"]):
        phase = CATEGORY_SECTION.get(r["Category"])
        if phase is None:
            unmapped.add(r["Category"])
            continue
        if r["Rank"] is None:
            continue
        rows.append({
            "phase": phase, "season": str(r["Season"]), "category": r["Category"], "stat": r["Stat"],
            "value": r["Value"], "rank": r["Rank"], "context": r["Context"], "week": None,
        })
    wb.close()

    finalized_seasons = {r["season"] for r in rows}
    weekly_files = []
    weeks_by_season = {}
    for path in sorted(NATIONAL_SRC.parent.glob("Carroll_*_Weekly.xlsx")):
        m = WEEKLY_YEAR_RE.search(path.name)
        season = m.group(1) if m else None
        if season is None or season in finalized_seasons:
            continue  # already superseded by that season's finalized AllYears rows
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Team Detail" not in wb.sheetnames:
            # Weekly file predates the Team Detail sheet (build_weekly_trends.ps1 fix,
            # 2026-07-30) -- team-level Rank isn't recoverable from this file, skip it
            # rather than showing an unranked/incomplete row.
            wb.close()
            continue
        season_weeks = {}
        for r in sheet_rows(wb["Team Detail"]):
            phase = CATEGORY_SECTION.get(r["Category"])
            if phase is None:
                unmapped.add(r["Category"])
                continue
            if r["Rank"] is None:
                continue
            rows.append({
                "phase": phase, "season": season, "category": r["Category"], "stat": r["Stat"],
                "value": r["Value"], "rank": r["Rank"], "context": r["Context"], "week": r["Week"],
            })
            season_weeks[r["Week"]] = str(r["SnapshotDate"])[:10]
        wb.close()
        if season_weeks:
            weeks_by_season[season] = [{"week": w, "date": season_weeks[w]} for w in sorted(season_weeks)]
            weekly_files.append(path.name)

    if unmapped:
        print(f"WARNING: {len(unmapped)} National category(ies) with no phase mapping, skipped: {sorted(unmapped)}")
    return rows, weekly_files, weeks_by_season


def main():
    cciw_rows, cciw_weekly_files, cciw_weeks = build_cciw()
    national_rows, national_weekly_files, national_weeks = build_national()

    def distinct(rows, field):
        return sorted({r[field] for r in rows if r.get(field) not in (None, "")}, key=str)

    payload = {
        "generated_from": {
            "cciw": [CCIW_SRC.name] + cciw_weekly_files,
            "national": [NATIONAL_SRC.name] + national_weekly_files,
        },
        "cciw": {
            "rows": cciw_rows,
            "seasons": distinct(cciw_rows, "season"),
            "weeks": cciw_weeks,
        },
        "national": {
            "rows": national_rows,
            "seasons": distinct(national_rows, "season"),
            "weeks": national_weeks,
        },
    }

    OUT.parent.mkdir(exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=0, default=str)

    if cciw_weekly_files or national_weekly_files:
        print(f"in-season weekly data included from: cciw={cciw_weekly_files} national={national_weekly_files}")

    for phase in ("offense", "defense", "special_teams", "overall"):
        c = sum(1 for r in cciw_rows if r["phase"] == phase)
        n = sum(1 for r in national_rows if r["phase"] == phase)
        print(f"{phase}: cciw={c} national={n}")


if __name__ == "__main__":
    main()
