"""
Reads combined_play_data.xlsx (sibling "Game Analysis" project) and writes
../data/game-data.json -- row-level Offense/Defense records for Phase 3
(offense.html / defense.html), same "row-level JSON, filter/aggregate
client-side" pattern as build_special_teams_data.py.

Two source sheets, two different grains:
  - Plays: Carroll's own hand-charted play-by-play (formation, personnel,
    play call, opponent front/blitz/coverage shown). No QUARTER column, no
    drive/score context. ODK marks whose snap this is: 'O' (Carroll offense),
    'D' (Carroll defense), 'K' (kicking -- already covered by the Special
    Teams Data pipeline), 'S'/blank (penalty/timeout/admin rows with no real
    snap side -- excluded from the offense/defense split entirely, not
    treated as a 3rd side).
  - OfficialPlayByPlay: the scraped official box-score log. Coarser play
    detail (no formation/personnel) but has quarter, drive number/result,
    score differential, explosive/situational efficiency already computed.
    Used for the Executive Scorecard tab; Plays is used for the Play-Calling
    & Tendencies tab.

Five raw files are scout games Carroll isn't even in (e.g. "Aurora vs SNC") --
GameMatchLog.HAS_OFFICIAL_PBP is False for those and they're dropped entirely,
same as the "no Carroll in filename" reason already recorded there. Only the
50 real Carroll games (2021-2025) are used.

Column semantics are per Game Analysis/.../DataDictionary sheet in the source
workbook -- read that (or this script's comments) before changing what any
field here means.

Re-run whenever the source workbook changes:
    python build_game_data.py
"""

import json
import re
from pathlib import Path

import openpyxl

SRC = Path(__file__).resolve().parent.parent.parent / "Game Analysis" / "processed" / "combined_play_data.xlsx"
OUT = Path(__file__).resolve().parent.parent / "data" / "game-data.json"

CARROLL = "Carroll (WI)"

# The official box-score scrape used a different OPPONENT string for the same
# school in different seasons -- confirmed by cross-referencing GAME_LABEL
# (which stays consistent, e.g. always "Carroll vs Wash U ...") against
# OPPONENT (which doesn't), and by season exclusivity (each pair never
# co-occurs in the same season, consistent with a one-time scrape naming
# change rather than two real different opponents): "Washington (Mo.)"
# (2021 only) / "WashU" (2022-2025), and "Wisconsin Lutheran" (2022 only) /
# "Wis. Lutheran" (2023 only). Canonicalized to the more common of each
# pair's two spellings. Found 2026-08-04 per the user noticing it directly
# on the Opponent Scouting page's opponent dropdown.
OPPONENT_ALIASES = {
    "Washington (Mo.)": "WashU",
    "Wis. Lutheran": "Wisconsin Lutheran",
}


def canonical_opponent(name):
    return OPPONENT_ALIASES.get(name, name)


def sheet_rows(ws):
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for r in ws.iter_rows(min_row=2, values_only=True):
        yield dict(zip(headers, r))


def parse_label_date(label):
    """Year/season from the GAME_LABEL filename-stem date suffix (e.g.
    'Carroll vs Augustana 10_18_25' -> 2025). Matches OfficialPlayByPlay's own
    GAME_DATE in 49/50 games -- the one mismatch (North Park 2025, 10/11 vs
    10/12) is a next-day discrepancy already flagged by the source's own
    MATCH_CONFIDENCE = 'Nearest date (+/-1 day)', not a parsing bug here."""
    m = re.search(r"(\d{1,2})_(\d{1,2})_(\d{2})$", label)
    if not m:
        return None
    return 2000 + int(m.group(3))


def normalize_team(name):
    """Strips the official scrape's incidental prefixes/suffixes (AP-poll
    rank, and a home/away venue-letter quirk seen on a few rows) so the
    remaining string can be compared directly against POSSESSION_TEAM/
    OPPONENT-style team names elsewhere in the sheet."""
    if not name:
        return None
    name = re.sub(r"^#\d+\s+", "", str(name))
    name = re.sub(r"\s+at\s+[A-Z]$", "", name)
    return name.strip()


def classify_side(possession_team, opponent):
    """'offense' if Carroll has the ball, 'defense' if the known opponent
    does, else None (administrative rows like 'Halftime'/'Game Start'/
    'Start of Quarter #N' -- not a real snap by either side, excluded).
    `opponent` is always the already-canonicalized OPPONENT value (see
    canonical_opponent()/OPPONENT_ALIASES) -- possession_team needs the same
    canonicalization applied before comparing, or every defensive snap in a
    game whose OPPONENT got renamed would stop matching its own
    POSSESSION_TEAM value (a real bug hit once building this: renaming
    "Washington (Mo.)" -> "WashU" silently dropped that entire game's
    defense-side rows, since POSSESSION_TEAM still said the old name when
    Carroll's opponent had the ball)."""
    norm = canonical_opponent(normalize_team(possession_team))
    if norm == CARROLL:
        return "offense"
    if norm == opponent:
        return "defense"
    return None


SCRIMMAGE_PLAY_TYPES = {"Rush", "Pass", "Sack", "Kneel", "Two-Point Conversion"}


def outcome_flags(outcome):
    tags = set((outcome or "").split(", ")) if outcome else set()
    return {
        "is_touchdown": "Touchdown" in tags,
        "is_penalty": "Penalty" in tags,
        "is_first_down": "First Down" in tags,
    }


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

    gml = list(sheet_rows(wb["GameMatchLog"]))
    carroll_game_labels = {r["GAME_LABEL"] for r in gml if r["HAS_OFFICIAL_PBP"]}

    opbp_all = list(sheet_rows(wb["OfficialPlayByPlay"]))
    for r in opbp_all:
        r["OPPONENT"] = canonical_opponent(r["OPPONENT"])
    opponent_by_game = {r["GAME_LABEL"]: r["OPPONENT"] for r in opbp_all}
    date_by_game = {r["GAME_LABEL"]: r["GAME_DATE"] for r in opbp_all}

    games = [
        {"game_label": gl, "season": parse_label_date(gl), "opponent": opp, "date": date_by_game.get(gl)}
        for gl, opp in opponent_by_game.items()
    ]
    games.sort(key=lambda g: (g["season"], g["date"] or ""))

    # ---------------------------------------------------------- Plays sheet --
    plays_all = [
        r for r in sheet_rows(wb["Plays"])
        if r["GAME_LABEL"] in carroll_game_labels and r["ODK"] in ("O", "D")
    ]

    offense_plays, defense_plays = [], []
    for r in plays_all:
        season = parse_label_date(r["GAME_LABEL"])
        opponent = opponent_by_game.get(r["GAME_LABEL"])
        base = {
            "season": season, "opponent": opponent,
            "game_label": r["GAME_LABEL"], "date": date_by_game.get(r["GAME_LABEL"]),
            "down": r["DN"], "distance": r["DIST"], "hash": r["HASH"],
            "field_pos": r["YARD LN"], "yards": r["GN/LS"],
            "play_type": r["PLAY_TYPE"], "play_outcome": r["PLAY_OUTCOME"],
            "turnover_type": r["TURNOVER_TYPE"], "play_efficiency": r["PLAY_EFFICIENCY"],
            "situation": r["SITUATION"], "field_zone": r["FIELD_ZONE"],
        }
        if r["ODK"] == "O":
            offense_plays.append({
                **base,
                "personnel": r["PERSONNEL"], "formation": r["OFF FORM"],
                "formation_strength": r["OFF STR"], "backfield": r["BACKFIELD"],
                "motion": r["MOTION"], "protection": r["PROTECTION"],
                "play_call": r["PLAY_CALL"],
                "def_front": r["DEF FRONT"], "stunt": r["STUNT"], "blitz": r["BLITZ"],
                "cov_shell": r["COV SHELL"], "coverage": r["COVERAGE"],
            })
        else:
            # NOTE: unlike 'O' rows (which also carry the opponent's DEF FRONT/
            # STUNT/BLITZ/COV SHELL/COVERAGE shown against Carroll's O), 'D' rows
            # do NOT carry the opponent's own OFF FORM/PERSONNEL/BACKFIELD/etc --
            # confirmed by checking real fill rates on the 50-real-game set
            # (PERSONNEL/OFF FORM/... are <1% filled on 'D' rows, vs 90%+ on 'O'
            # rows). This is a structural property of the self-scout charting
            # workflow (own-team detail + opponent's D shown to the O), not a
            # "still being charted" gap like Snap Location -- so there's no
            # opp_formation/opp_personnel/etc field here at all, and the Defense
            # dashboard should say so explicitly rather than build unused UI for it.
            defense_plays.append({
                **base,
                "front_d": r["FRONT (D)"], "tag_d": r["TAG (D)"], "movement": r["MOVEMENT"],
                "blitz_d": r["BLITZ (D)"], "num_d": r["#(D)"],
            })

    # ------------------------------------------------- OfficialPlayByPlay ----
    offense_official, defense_official = [], []
    drives_by_key = {}  # (game_label, drive_num) -> {side, season, opponent, ...}

    for r in opbp_all:
        if r["GAME_LABEL"] not in carroll_game_labels:
            continue
        side = classify_side(r["POSSESSION_TEAM"], r["OPPONENT"])
        if side is None:
            continue
        season = parse_label_date(r["GAME_LABEL"])

        if r["DRIVE_NUM"] is not None:
            key = (r["GAME_LABEL"], r["DRIVE_NUM"])
            if key not in drives_by_key:
                drives_by_key[key] = {
                    "season": season, "opponent": r["OPPONENT"], "date": r["GAME_DATE"], "side": side,
                    "start_field_pos": r["DRIVE_START_FIELD_POS"],
                    "play_count": r["DRIVE_PLAY_COUNT"], "yards": r["DRIVE_YARDS"],
                    "result": r["DRIVE_RESULT"],
                }

        if r["PLAY_TYPE"] not in SCRIMMAGE_PLAY_TYPES:
            continue
        row = {
            "season": season, "opponent": r["OPPONENT"],
            "game_label": r["GAME_LABEL"], "date": r["GAME_DATE"], "quarter": r["QUARTER"],
            "down": r["DOWN"], "distance": r["DISTANCE"], "is_goal_to_go": r["IS_GOAL_TO_GO"],
            "field_zone": r["FIELD_ZONE"], "situation": r["SITUATION"],
            "play_type": r["PLAY_TYPE"], "direction": r["DIRECTION"], "pass_depth": r["PASS_DEPTH"],
            "yards": r["YARDS"], "play_efficiency": r["PLAY_EFFICIENCY"],
            "play_outcome": r["PLAY_OUTCOME"], "is_turnover": r["IS_TURNOVER"],
            "score_differential": r["SCORE_DIFFERENTIAL"],
            # Broadcast from the drive this play belongs to (same convention the
            # source workbook already uses for DRIVE_YARDS etc) so a Red Zone /
            # drive-result join works directly off a single play row, without
            # also having to match back into the separate `drives` array below.
            "drive_num": r["DRIVE_NUM"], "drive_result": r["DRIVE_RESULT"],
            **outcome_flags(r["PLAY_OUTCOME"]),
        }
        (offense_official if side == "offense" else defense_official).append(row)

    offense_drives = [{"drive_num": k[1], "game_label": k[0], **v} for k, v in drives_by_key.items() if v["side"] == "offense"]
    defense_drives = [{"drive_num": k[1], "game_label": k[0], **v} for k, v in drives_by_key.items() if v["side"] == "defense"]
    for d in (*offense_drives, *defense_drives):
        del d["side"]

    def distinct(rows, field):
        return sorted({r[field] for r in rows if r.get(field) not in (None, "")}, key=str)

    payload = {
        "generated_from": SRC.name,
        "games": games,
        "offense": {"plays": offense_plays, "official": offense_official, "drives": offense_drives},
        "defense": {"plays": defense_plays, "official": defense_official, "drives": defense_drives},
        "filters": {
            "offense": {
                "season": distinct(offense_plays, "season"), "opponent": distinct(offense_plays, "opponent"),
                "situation": distinct(offense_plays, "situation"), "field_zone": distinct(offense_plays, "field_zone"),
                "down": distinct(offense_plays, "down"), "hash": distinct(offense_plays, "hash"),
                "personnel": distinct(offense_plays, "personnel"), "formation": distinct(offense_plays, "formation"),
                "play_call": distinct(offense_plays, "play_call"), "def_front": distinct(offense_plays, "def_front"),
                "coverage": distinct(offense_plays, "coverage"), "play_type": distinct(offense_plays, "play_type"),
                "quarter": distinct(offense_official, "quarter"), "direction": distinct(offense_official, "direction"),
            },
            "defense": {
                "season": distinct(defense_plays, "season"), "opponent": distinct(defense_plays, "opponent"),
                "situation": distinct(defense_plays, "situation"), "field_zone": distinct(defense_plays, "field_zone"),
                "down": distinct(defense_plays, "down"), "hash": distinct(defense_plays, "hash"),
                "front_d": distinct(defense_plays, "front_d"), "blitz_d": distinct(defense_plays, "blitz_d"),
                "play_type": distinct(defense_plays, "play_type"),
                "quarter": distinct(defense_official, "quarter"), "direction": distinct(defense_official, "direction"),
            },
        },
    }

    OUT.parent.mkdir(exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=0, default=str)

    print(f"games: {len(games)}")
    print(f"offense: plays={len(offense_plays)} official={len(offense_official)} drives={len(offense_drives)}")
    print(f"defense: plays={len(defense_plays)} official={len(defense_official)} drives={len(defense_drives)}")
    print(f"wrote {OUT} ({OUT.stat().st_size / 1024:.0f} KB)")


if __name__ == "__main__":
    main()
