"""
Reads Carroll_Special_Teams_2021_Current.xlsx (sibling "Special Teams Data" project)
and writes ../data/special-teams.json -- row-level records for all 5 units (Punt,
Punt Return, Kickoff, Kickoff Return, PAT/FG "Money Unit"), enriched with the
derived fields the dashboard's charts need (field-position bucket, points scored,
etc.). Filtering/aggregation happens client-side in the dashboard's JS against this
row-level data, same pattern as Special Teams Data/mockups/*.html.

Column semantics are per Special Teams Data/rules/*.md -- read those before changing
what any field here means, not just this script's comments.

Re-run whenever the source workbook changes:
    python build_special_teams_data.py
"""

import json
from pathlib import Path

import openpyxl

SRC = Path(__file__).resolve().parent.parent.parent / "Special Teams Data" / "Carroll_Special_Teams_2021_Current.xlsx"
OUT = Path(__file__).resolve().parent.parent / "data" / "special-teams.json"


def sheet_rows(ws):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    for r in range(2, ws.max_row + 1):
        yield dict(zip(headers, (ws.cell(r, c).value for c in range(1, ws.max_column + 1))))


def field_bucket(converted_los):
    """10-yard buckets on the 0-100 Converted LOS scale, e.g. 27 -> '21-30'.
    Matches the bucketing shown in every 'Drive Success by Field Position'-style
    chart in the reference screenshots."""
    if converted_los is None:
        return None
    lo = max(0, min(90, (int(converted_los) // 10) * 10))
    return f"{lo + 1 if lo else 0}-{lo + 10}"


def drive_success(points):
    """A drive that scored (any points > 0) counts as a 'success' for the
    Drive-Success-rate charts -- matches the reference screenshots' % bars."""
    if points is None:
        return None
    return points > 0


def outcome_bucket(points):
    """No Score / Field Goal / Touchdown, matching the 3-way stacked-bar legend in
    the reference screenshots. Safety (2, against Carroll) shown as its own bucket
    only where a sheet's real data has it, so it isn't silently dropped."""
    if points is None:
        return None
    if points == 0:
        return "No Score"
    if points == 3:
        return "Field Goal"
    if points == 2:
        return "Safety"
    if points in (6, 7, 8):
        return "Touchdown"
    return f"{points} pts"


def build_punt(wb):
    ws = wb["Carroll Punt"]
    rows = []
    for r in sheet_rows(ws):
        pts = r["Next Drive Outcome"]
        snap_to_kick = None
        if r["Snap to Catch"] is not None and r["Catch to Kick"] is not None:
            snap_to_kick = round(r["Snap to Catch"] + r["Catch to Kick"], 3)
        rows.append({
            "season": r["Season"], "date": r["Date"], "opponent": r["Opponent"],
            "is_home": r["Is Home"], "quarter": r["Quarter"],
            "punter": r["Punter Name"], "snapper": r["Snapper Name"],
            "kick_outcome": r["Kick Outcome"], "next_drive_points": pts,
            "next_drive_outcome": outcome_bucket(pts), "drive_success": drive_success(pts),
            "hangtime": r["Hangtime"], "total_distance": r["Total Distance"],
            "return_length": r["Return Length"], "snap_to_kick": snap_to_kick,
            "converted_los": r["Converted Result LOS"],
            "field_bucket": field_bucket(r["Converted Result LOS"]),
            "i20": r["I20?"],
        })
    return rows


def punt_return_outcome(r):
    """Carroll Punt Return has no single 'outcome' column (unlike Carroll Punt) --
    derive one from the boolean flags, in the same priority order the sheet's own
    rule-based overrides use (a play can't be more than one of these)."""
    if r["TB"]:
        return "Touchback"
    if r["Blocked?"]:
        return "Blocked"
    if r["Muff?"]:
        return "Muff"
    if r["FC"]:
        return "Fair Catch"
    if (r["Return Length"] or 0) > 0:
        return "Returned"
    return "Downed"


def build_punt_return(wb):
    ws = wb["Carroll Punt Return"]
    rows = []
    for r in sheet_rows(ws):
        pts = r["Next Drive Outcome"]
        snap_to_kick = None
        if r["Snap to Catch"] is not None and r["Catch to Kick"] is not None:
            snap_to_kick = round(r["Snap to Catch"] + r["Catch to Kick"], 3)
        rows.append({
            "season": r["Season"], "date": r["Date"], "opponent": r["Opponent"],
            "is_home": r["Is Home"], "quarter": r["Quarter"], "returner": r["Returner"],
            "kick_outcome": punt_return_outcome(r), "next_drive_points": pts,
            "next_drive_outcome": outcome_bucket(pts), "drive_success": drive_success(pts),
            "hangtime": r["Hangtime"], "carry_distance": r["Total Distance"],
            "return_length": r["Return Length"], "snap_to_kick": snap_to_kick,
            "converted_los": r["Converted Result LOS"],
            "field_bucket": field_bucket(r["Converted Result LOS"]),
        })
    return rows


def build_kickoff(wb):
    ws = wb["Carroll Kickoff"]
    rows = []
    for r in sheet_rows(ws):
        pts = r["Next Drive Outcome"]
        rows.append({
            "season": r["Season"], "date": r["Date"], "opponent": r["Opponent"],
            "is_home": r["Is Home"], "quarter": r["Quarter"], "kicker": r["Kicker"],
            "kick_type": r["Kick Type"], "next_drive_points": pts,
            "next_drive_outcome": outcome_bucket(pts), "drive_success": drive_success(pts),
            "hangtime": r["Hangtime"], "total_distance": r["Total Distance"],
            "return_length": r["Return Length"], "hash_kicked_from": r["Hash Kicked From"],
            "converted_los": r["Converted Result LOS"],
            "field_bucket": field_bucket(r["Converted Result LOS"]),
        })
    return rows


def build_kickoff_return(wb):
    ws = wb["Carroll Kickoff Return"]
    rows = []
    for r in sheet_rows(ws):
        pts = r["Next Drive Outcome"]
        rows.append({
            "season": r["Season"], "date": r["Date"], "opponent": r["Opponent"],
            "is_home": r["Is Home"], "quarter": r["Quarter"], "returner": r["Returner"],
            "kick_type": r["Kick Type"], "next_drive_points": pts,
            "next_drive_outcome": outcome_bucket(pts), "drive_success": drive_success(pts),
            "hangtime": r["Hangtime"], "total_distance": r["Total Distance"],
            "return_length": r["Return Length"],
            "return_location": r["Hash Received On"],
            "converted_los": r["Converted Result LOS"],
            "field_bucket": field_bucket(r["Converted Result LOS"]),
        })
    return rows


def build_money_unit(wb):
    ws = wb["Carroll PAT-FG"]
    rows = []
    for r in sheet_rows(ws):
        snap_to_kick = None
        if r["Snap to Catch"] is not None and r["Catch to Kick"] is not None:
            snap_to_kick = round(r["Snap to Catch"] + r["Catch to Kick"], 3)
        rows.append({
            "season": r["Season"], "date": r["Date"], "opponent": r["Opponent"],
            "is_home": r["Is Home"], "quarter": r["Quarter"], "kicker": r["Kicker"],
            "long_snapper": r["Long Snapper"], "fg_exp": r["FG/EXP"],
            "distance": r["Distance"], "make": r["Make?"],
            "miss_location": r["Miss Location"],
            "hash_kicked_from": r["Hash Kicked From"],
            "snap_to_catch": r["Snap to Catch"], "catch_to_kick": r["Catch to Kick"],
            "snap_to_kick": snap_to_kick,
            "value": r["PAT/FG Value"], "score": r["PAT/FG Score"],
        })
    return rows


def distinct(rows, field):
    # Sort by string form -- some fields mix types (quarter is int 1-4 but "OT" is
    # a str), and Python can't compare int/str directly (crashes sorted() on any
    # unit that has an OT row, which money_unit does).
    return sorted({r[field] for r in rows if r.get(field) not in (None, "")}, key=str)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    punt = build_punt(wb)
    punt_return = build_punt_return(wb)
    kickoff = build_kickoff(wb)
    kickoff_return = build_kickoff_return(wb)
    money = build_money_unit(wb)

    payload = {
        "generated_from": SRC.name,
        "units": {
            "punt": punt,
            "punt_return": punt_return,
            "kickoff": kickoff,
            "kickoff_return": kickoff_return,
            "money_unit": money,
        },
        "filters": {
            "punt": {
                "season": distinct(punt, "season"), "opponent": distinct(punt, "opponent"),
                "punter": distinct(punt, "punter"), "snapper": distinct(punt, "snapper"),
                "next_drive_outcome": distinct(punt, "next_drive_outcome"),
            },
            "punt_return": {
                "season": distinct(punt_return, "season"), "opponent": distinct(punt_return, "opponent"),
                "returner": distinct(punt_return, "returner"),
                "next_drive_outcome": distinct(punt_return, "next_drive_outcome"),
                "kick_outcome": distinct(punt_return, "kick_outcome"),
            },
            "kickoff": {
                "season": distinct(kickoff, "season"), "opponent": distinct(kickoff, "opponent"),
                "kicker": distinct(kickoff, "kicker"), "kick_type": distinct(kickoff, "kick_type"),
                "next_drive_outcome": distinct(kickoff, "next_drive_outcome"),
            },
            "kickoff_return": {
                "season": distinct(kickoff_return, "season"), "opponent": distinct(kickoff_return, "opponent"),
                "returner": distinct(kickoff_return, "returner"), "kick_type": distinct(kickoff_return, "kick_type"),
                "next_drive_outcome": distinct(kickoff_return, "next_drive_outcome"),
            },
            # return_location intentionally not offered as a *filter* (it's a chart
            # dimension, matching the reference dashboard, which doesn't filter on it
            # either) -- distinct values available in the data if that changes:
            # L / LM / M / RM / R / OBR / OBL.
            "money_unit": {
                "season": distinct(money, "season"), "kicker": distinct(money, "kicker"),
                "long_snapper": distinct(money, "long_snapper"), "fg_exp": distinct(money, "fg_exp"),
            },
        },
    }

    # quarter/is_home apply to every unit -- compute from each unit's own rows
    # rather than hardcoding ["1","2","3","4"] on the dashboard side. Real bug
    # found 2026-07-29 testing in the browser: 3 Money Unit rows have
    # quarter == "OT" (overtime), which a hardcoded 1-4 list silently excluded
    # from the default (everything-selected) filter view with no error -- the
    # rows just vanished. Computing this from the data means a future OT/2OT
    # row in any unit shows up as a real filter option automatically.
    for unit, rows in payload["units"].items():
        payload["filters"][unit]["quarter"] = distinct(rows, "quarter")
        payload["filters"][unit]["is_home"] = distinct(rows, "is_home")

    OUT.parent.mkdir(exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=0, default=str)

    for unit, rows in payload["units"].items():
        print(f"{unit}: {len(rows)} rows")


if __name__ == "__main__":
    main()
